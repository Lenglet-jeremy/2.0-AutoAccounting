from openpyxl import *
from random import randint
import datetime

ENTREPRISE_NAME = "Entreprise fictive";
EMPLOYEES = 5000;
TEMP_WORK = randint(round(EMPLOYEES * 0.07), round(EMPLOYEES * 0.12));
TEMP_WORK_5_LAST_YEARS = TEMP_WORK * 12 * 5;
print(TEMP_WORK)

CurrentDay = datetime.date.today()

def lastDayOfPreviousMounth(date) : 
    if(str(date).split("-")):
        date = date.replace(day = 1)
        last_month = date - datetime.timedelta(days=1)
        return last_month

lastDayOfPreviousMounth(CurrentDay)

COST_CENTER = {
    "6200" : "Usine",
    "6201" : "Montage",
    "6202" : "Usinage",
    "6203" : "Magasin",
    "6204" : "Expédition",
    "6205" : "Reception",
    "6206" : "Restauration",
    "6207" : "Propreté",
    "6208" : "Logistique"
}

PROFIT_CENTER = {
    "7200" : "Usine",
    "7201" : "Montage",
    "7202" : "Usinage",
    "7203" : "Magasin",
    "7204" : "Expédition",
    "7205" : "Reception",
    "7206" : "Restauration",
    "7207" : "Propreté",
    "7208" : "Logistique"
}

income_statement_DB_file = Workbook();
DB_ws = income_statement_DB_file.active;
DB_ws.title = "DB";

DB_ws["A1"] = "Nature comptable";
for i in range(2, (TEMP_WORK * 12 * 5) + 2): 
    DB_ws.cell(row = i, column = 1).value = "62110001";

DB_ws["B1"] = "Designation comptable";
for i in range(2, (TEMP_WORK * 12 * 5) + 2): 
    DB_ws.cell(row = i, column = 2).value = "Personnel intérimaire";


DB_ws["C1"] = "Centre de coût";
for i in range(2, TEMP_WORK_5_LAST_YEARS + 2): 
    DB_ws.cell(row = i, column = 3).value = randint(6200, 6208);

DB_ws["D1"] = "Designation centre de coût";
# # Saisie des désignations des centres de coûts des 5 dernières années
# # En fonction des centres de coûts
a = 0
for i in range(2, (TEMP_WORK * 12 * 5) + 2): 
    for ii in range(1, len(COST_CENTER)) :

        cellValue = str(DB_ws.cell(row = i, column = 3).value);

        if(COST_CENTER[cellValue]) : 
            DB_ws.cell(row = i, column = 4).value = COST_CENTER[cellValue];
        


DB_ws["E1"] = "Centre de profit";
DB_ws["F1"] = "Designation centre de profit";

DB_ws["G1"] = "Montant";
for i in range(2, TEMP_WORK_5_LAST_YEARS + 2): 
    DB_ws.cell(row = i, column = 7).value = randint(4000, 5000);
DB_ws["H1"] = "Type Piece";
for i in range(2, TEMP_WORK_5_LAST_YEARS + 2): 
    DB_ws.cell(row = i, column = 8).value = "Facture";


DB_ws["I1"] = "Nom";
DB_ws["J1"] = "Prenom";
DB_ws["K1"] = "Matricule";
for i in range(2, TEMP_WORK_5_LAST_YEARS + 2): 
    NbTempWork = randint(1, TEMP_WORK)
    DB_ws.cell(row = i, column = 9).value = "nom" + str(NbTempWork);
    DB_ws.cell(row = i, column = 10).value = "prenom" + str(NbTempWork);
    DB_ws.cell(row = i, column = 11).value = 100000 + int(NbTempWork);

DB_ws["L1"] = "Periode d'effet";
DB_ws["M1"] = "Debut periode";
DB_ws["N1"] = "Fin periode";

DB_ws["O1"] = "N° piece reference";

DB_ws["P1"] = "Utilisateur ecriture";

DB_ws["Q1"] = "Date piece";
DB_ws["R1"] = "Date comptable";
DB_ws["S1"] = "Date de saisie";

DB_ws["T1"] = "Compte contre partie";
DB_ws["U1"] = "Designation compte contre partie";

DB_ws["V1"] = "N° Ecriture";

DB_ws["W1"] = "Commentaire ecriture";

DB_ws["X1"] = "N° contre passation";
DB_ws["Y1"] = "Commentaire contre passation";

DB_ws["Z1"] = "Devise";
DB_ws["AA1"] = "Convertion en euros";
DB_ws["AB1"] = "Date convertion";
DB_ws["AC1"] = "Taux convertion";
DB_ws["AD1"] = "Source convertion";

DB_ws["AE1"] = "Societe";
DB_ws["AF1"] = "Designation societe";

DB_ws["AG1"] = "Unité de quantité";

DB_ws["AH1"] = "Quantité";
DB_ws["AI1"] = "Taux unité de quantité";

DB_ws["AJ1"] = "Code mouvement";
DB_ws["AK1"] = "Designation mouvement";


def cellToNums(cell) :
    pass

def numsToCell(row, col) :
    pass


def rangeToNums(range) : 
    pass

def numsToRange(startRow, startCol, endRow, endCol) : 
    pass



income_statement_DB_file.save("./0.0.Template/income_statement_DB_file.xlsx");